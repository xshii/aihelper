"""Excel → Table importer.

每个 sheet 生成一个 ``Table(base_name=<sheet PascalCase>)``：首列（或 ``index_overrides``
指定列）进 ``Record.index``，其余列进 ``Record.attribute``；不产出 ``ref`` —— legacy
Excel 表不表达关联关系，由 schema 或 yaml 编辑层补充。
"""
from __future__ import annotations

import re
from contextlib import closing
from pathlib import Path
from typing import Any, Dict, FrozenSet, Iterator, List, NamedTuple, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ecfg.model import CellValue, Record, Table

# region name normalization ────────────────────────────────────────────────
_NON_WORD = re.compile(r"[^A-Za-z0-9]+")
_CAMEL_BOUNDARY = re.compile(r"(?<=[a-z0-9])(?=[A-Z])")


def _split_words(s: str) -> List[str]:
    """把 'IRQ_TABLE' / 'core id' / 'irqTbl' 这类串拆成词列表。"""
    s = s.strip()
    if not s:
        return []
    out: List[str] = []
    for chunk in _NON_WORD.split(s):
        if chunk:
            out.extend(w for w in _CAMEL_BOUNDARY.split(chunk) if w)
    return out


def _capitalize(word: str) -> str:
    """首字母大写，其余小写；保留 'IRQ' → 'Irq' 这类缩写的可读性。"""
    return word[0].upper() + word[1:].lower()


def to_pascal_case(s: str) -> str:
    """任意书写风格 → PascalCase，如 'IRQ_TABLE' → 'IrqTable'。"""
    words = _split_words(s)
    if not words:
        raise ValueError(f"cannot derive PascalCase from {s!r}")
    return "".join(_capitalize(w) for w in words)


def to_camel_case(s: str) -> str:
    """任意书写风格 → camelCase，如 'core id' → 'coreId'。"""
    pascal = to_pascal_case(s)
    return pascal[0].lower() + pascal[1:]


# endregion

# region value coercion ─────────────────────────────────────────────────────
def _cell_value(v: Any) -> CellValue:
    """openpyxl cell → YAML 友好的 Python 值。空串/纯空白 → None。"""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        stripped = v.strip()
        return stripped or None
    return str(v)


# endregion

# region value objects ──────────────────────────────────────────────────────
class Header(NamedTuple):
    """列表头的原名（Excel 显示）和归一化后的字段名（YAML 输出）。"""

    raw: str
    name: str


# endregion

# region sheet → records ────────────────────────────────────────────────────
class SheetConverter:
    """把一张 worksheet 转成 ``Table``。"""

    def __init__(self, sheet: Worksheet, index_overrides: Dict[str, List[str]]):
        """预计算 base_name / 表头 / index 列集合，后续方法共享这些缓存。"""
        self._sheet = sheet
        self._overrides = index_overrides
        self.base_name: str = to_pascal_case(sheet.title)
        self._headers: List[Header] = list(self._iter_headers())
        self._index_names: FrozenSet[str] = frozenset(self._resolve_index_cols())

    def convert(self) -> List[Record]:
        """迭代数据行并产出 Record 列表；全空行自动跳过。"""
        records: List[Record] = []
        for row in self._sheet.iter_rows(min_row=2, values_only=True):
            rec = self._row_to_record(row)
            if rec is not None:
                records.append(rec)
        return records

    def _iter_headers(self) -> Iterator[Header]:
        """从第 1 行产出 Header；表头空或字段名重名直接 raise。"""
        row = next(self._sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if row is None:
            raise ValueError(f"sheet {self._sheet.title!r} 为空")
        seen: Set[str] = set()
        for col_idx, raw in enumerate(row, start=1):
            if raw is None or str(raw).strip() == "":
                raise ValueError(
                    f"sheet {self._sheet.title!r}: 第 {col_idx} 列表头为空"
                )
            name = to_camel_case(str(raw))
            if name in seen:
                raise ValueError(
                    f"sheet {self._sheet.title!r}: 字段名 {name!r} 重复（原 {raw!r}）"
                )
            seen.add(name)
            yield Header(raw=str(raw), name=name)

    def _resolve_index_cols(self) -> List[str]:
        """决定哪些列作 ``index``：用户覆盖优先（按 sheet title 或 base_name 匹配），否则首列。"""
        for key in (self._sheet.title, self.base_name):
            if key in self._overrides:
                return [self._match_col(c) for c in self._overrides[key]]
        return [self._headers[0].name]

    def _match_col(self, col: str) -> str:
        """把用户输入的列名（原名/camelCase/任意风格）匹配到归一化的字段名。"""
        cam = to_camel_case(col)
        for h in self._headers:
            if col == h.raw or col == h.name or cam == h.name:
                return h.name
        raise ValueError(
            f"sheet {self._sheet.title!r}: 找不到 index 列 {col!r}"
        )

    def _row_to_record(self, row_values: Tuple[Any, ...]) -> Optional[Record]:
        """把一行单元格值拆成 Record；全空返回 None 以便跳过。"""
        index: Dict[str, CellValue] = {}
        attribute: Dict[str, CellValue] = {}
        has_any = False
        for header, val in zip(self._headers, row_values):
            cooked = _cell_value(val)
            if cooked is not None:
                has_any = True
            target = index if header.name in self._index_names else attribute
            target[header.name] = cooked
        if not has_any:
            return None
        return Record(index=index, attribute=attribute)


# endregion

# region public API ─────────────────────────────────────────────────────────
def excel_to_tables(
    xlsx_path: Path,
    *,
    index_overrides: Optional[Dict[str, List[str]]] = None,
) -> List[Table]:
    """读 xlsx，每个非空 sheet 产出一个 Table；跳过空 sheet 和 <2 行的 sheet。"""
    overrides = index_overrides or {}
    tables: List[Table] = []
    with closing(load_workbook(xlsx_path, data_only=True)) as wb:
        for sheet in wb.worksheets:
            if sheet.max_row is None or sheet.max_row < 2:
                continue
            conv = SheetConverter(sheet, overrides)
            records = conv.convert()
            if not records:
                continue
            tables.append(Table(
                base_name=conv.base_name,
                records=records,
                source_hint=xlsx_path.name,
            ))
    return tables


# endregion
